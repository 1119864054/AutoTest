import pypinyin
from openpyxl import load_workbook
from pypinyin import Style

test_dcit = {}
test_list = []


def trans(value):
    result = pypinyin.slug(value, style=Style.NORMAL, strict=False, separator='')
    test_dcit[value] = result
    test_list.append(result)
    return result


if __name__ == '__main__':
    wb = load_workbook('C:\\Users\\11198\\Desktop\\test_pinyin.xlsx')
    ws = wb['Sheet1']
    for row in range(len(list(ws.rows))):
        res = trans(ws.cell(row=row+1, column=1).value)
        ws.cell(row=row+1, column=2).value = res
    wb.save('C:\\Users\\11198\\Desktop\\test_pinyin.xlsx')
    print(test_dcit)
    print(test_list)
